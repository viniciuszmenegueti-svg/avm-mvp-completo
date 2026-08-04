"""Reconcile two VivaReal workbooks and train a research-only OLS candidate.

The input files contain asking prices, not verified transaction/market values.
For that reason the resulting model is deliberately persisted as CANDIDATE and
the API blocks both homologation approval and AVM inference.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import math
import re
import unicodedata
from collections import Counter
from collections.abc import Iterable
from datetime import datetime, timezone
from pathlib import Path
from statistics import median
from typing import Any
from urllib.error import HTTPError, URLError
from urllib.request import Request, urlopen

from dotenv import dotenv_values
from openpyxl import load_workbook


CITY = "Rio de Janeiro"
STATE = "RJ"
CITY_IBGE_CODE = "3304557"
PROPERTY_TYPE = "APARTMENT"
FEATURE_MAPPING = (
    ("usableArea", "private_area_m2"),
    ("bedrooms", "bedrooms"),
    ("bathrooms", "bathrooms"),
    ("parkingSpaces", "parking_spaces"),
)
CORE_RECONCILIATION_FIELDS = (
    "url",
    "id",
    "externalId",
    "listingType",
    "propertyType",
    "city",
    "state",
    "currency",
    "price",
    "usableArea",
    "bedrooms",
    "bathrooms",
    "parkingSpaces",
    "street",
    "neighborhood",
    "zipCode",
    "lat",
    "lng",
    "updatedAt",
    "scrapedAt",
)
KNOWN_PLACEHOLDER_SENTINEL = "143"
PLACEHOLDER_TEXT_FIELDS = ("publicationType", "status")
MAX_PLACEHOLDER_SHARE = 0.50


def parse_arguments() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Concilia duas exportações VivaReal, preserva exclusões e registra "
            "um treino exploratório que não pode ser aprovado para AVM."
        )
    )
    parser.add_argument(
        "workbooks",
        nargs="+",
        type=Path,
        help=(
            "Uma ou mais exportações. Arquivos integralmente duplicados não "
            "aumentam a amostra e planilhas com sentinelas artificiais são rejeitadas."
        ),
    )
    parser.add_argument("--base-url", default="http://localhost:8001")
    parser.add_argument("--env-file", type=Path, default=Path(".env.homologation"))
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=Path(".audit") / "market-data" / "vivareal-rj",
    )
    parser.add_argument("--trainer-actor")
    parser.add_argument(
        "--skip-api",
        action="store_true",
        help="Gera apenas CSV e manifesto, sem registrar o candidato na API.",
    )
    return parser.parse_args()


def file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for block in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def load_rows(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if len(workbook.sheetnames) != 1:
            raise ValueError(f"{path.name}: esperado exatamente uma planilha.")
        sheet = workbook[workbook.sheetnames[0]]
        values = sheet.iter_rows(values_only=True)
        headers = [str(value or "").strip() for value in next(values)]
        if len(headers) != len(set(headers)):
            raise ValueError(f"{path.name}: existem cabeçalhos duplicados.")
        missing = sorted(set(CORE_RECONCILIATION_FIELDS) - set(headers))
        if missing:
            raise ValueError(
                f"{path.name}: campos obrigatórios ausentes: {', '.join(missing)}"
            )
        rows = [
            dict(zip(headers, row, strict=True))
            for row in values
            if any(value is not None and str(value).strip() for value in row)
        ]
        _reject_placeholder_contamination(path, rows)
        return rows
    finally:
        workbook.close()


def _reject_placeholder_contamination(path: Path, rows: list[dict[str, Any]]) -> None:
    """Reject exports where a known spreadsheet placeholder replaced nulls."""

    if not rows:
        return
    contaminated: list[str] = []
    for field in PLACEHOLDER_TEXT_FIELDS:
        matches = sum(
            _text(row.get(field)) == KNOWN_PLACEHOLDER_SENTINEL for row in rows
        )
        if matches / len(rows) >= MAX_PLACEHOLDER_SHARE:
            contaminated.append(f"{field}={matches}/{len(rows)}")
    if contaminated:
        details = ", ".join(contaminated)
        raise ValueError(
            f"{path.name}: possível preenchimento artificial pelo sentinela "
            f"{KNOWN_PLACEHOLDER_SENTINEL} ({details}). Preserve a fonte como "
            "evidência, mas não a use no treinamento."
        )


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _number(value: Any) -> float | None:
    text = _text(value).replace(" ", "")
    if not text:
        return None
    try:
        result = float(text)
    except ValueError:
        return None
    return result if math.isfinite(result) else None


def normalize_coordinate(value: Any, *, degree_digits: int) -> float | None:
    """Recover decimal coordinates exported with stripped/grouped separators."""

    text = _text(value).replace(" ", "")
    if not text:
        return None
    try:
        direct = float(text)
    except ValueError:
        direct = math.nan
    maximum = 90.0 if degree_digits == 2 else 180.0
    if math.isfinite(direct) and -maximum <= direct <= maximum:
        return direct
    sign = -1.0 if text.startswith("-") else 1.0
    digits = re.sub(r"\D", "", text)
    if len(digits) <= degree_digits:
        return None
    recovered = sign * int(digits) / (10 ** (len(digits) - degree_digits))
    return recovered if -maximum <= recovered <= maximum else None


def _normalized_text(value: Any) -> str:
    decomposed = unicodedata.normalize("NFKD", _text(value))
    ascii_text = "".join(char for char in decomposed if not unicodedata.combining(char))
    return " ".join(re.sub(r"[^A-Za-z0-9]+", " ", ascii_text).upper().split())


def _canonical_hash(value: object) -> str:
    serialized = json.dumps(
        value,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
        allow_nan=False,
    )
    return hashlib.sha256(serialized.encode("utf-8")).hexdigest()


def reconcile_workbooks(
    sources: list[tuple[Path, list[dict[str, Any]]]],
) -> tuple[list[dict[str, Any]], dict[str, int]]:
    by_url: dict[str, dict[str, Any]] = {}
    source_names: dict[str, list[str]] = {}
    conflicts: dict[str, int] = {field: 0 for field in CORE_RECONCILIATION_FIELDS}
    for path, rows in sources:
        for row in rows:
            url = _text(row.get("url"))
            if not url:
                url = f"MISSING-URL:{path.name}:{_text(row.get('id'))}"
            if url not in by_url:
                by_url[url] = dict(row)
                source_names[url] = [path.name]
                continue
            source_names[url].append(path.name)
            selected = by_url[url]
            for field in CORE_RECONCILIATION_FIELDS:
                left = _text(selected.get(field))
                right = _text(row.get(field))
                if left and right and field not in {"lat", "lng"} and left != right:
                    conflicts[field] += 1
                if not left and right:
                    selected[field] = row.get(field)
    rows = []
    for url in sorted(by_url):
        row = by_url[url]
        row["_source_files"] = "|".join(sorted(set(source_names[url])))
        rows.append(row)
    return rows, {key: value for key, value in conflicts.items() if value}


def prepare_audit_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    audit_rows: list[dict[str, Any]] = []
    for index, row in enumerate(rows, start=1):
        price = _number(row.get("price"))
        area = _number(row.get("usableArea"))
        bedrooms = _number(row.get("bedrooms"))
        bathrooms = _number(row.get("bathrooms"))
        parking = _number(row.get("parkingSpaces"))
        latitude = normalize_coordinate(row.get("lat"), degree_digits=2)
        longitude = normalize_coordinate(row.get("lng"), degree_digits=2)
        reasons: list[str] = []
        anomalies: list[str] = []
        if _text(row.get("listingType")).upper() != "SALE":
            reasons.append("NOT_SALE")
        if _normalized_text(row.get("city")) != "RIO DE JANEIRO":
            reasons.append("CITY_MISMATCH")
        if _text(row.get("state")).upper() != STATE:
            reasons.append("STATE_MISMATCH")
        if _text(row.get("propertyType")).upper() != "UNIT":
            reasons.append("PROPERTY_TYPE_MISMATCH")
        if _text(row.get("currency")).upper() != "BRL":
            reasons.append("CURRENCY_MISMATCH")
        values = (price, area, bedrooms, bathrooms, parking)
        if any(value is None for value in values):
            reasons.append("MISSING_TRAINING_FEATURE")
        elif price <= 0 or area <= 0 or min(bedrooms, bathrooms, parking) < 0:
            reasons.append("INVALID_NUMERIC_VALUE")
        elif area > 10_000 or max(bedrooms, bathrooms, parking) > 20:
            reasons.append("IMPLAUSIBLE_TRAINING_FEATURE")
        price_per_m2 = price / area if price and area and area > 0 else None
        if price_per_m2 is not None and not 5_000 <= price_per_m2 <= 30_000:
            anomalies.append("PRICE_PER_M2_REVIEW")
        if latitude is None or longitude is None:
            anomalies.append("COORDINATE_UNAVAILABLE_OR_INVALID")
        fingerprint = _canonical_hash(
            {
                "street": _normalized_text(row.get("street")),
                "postal_code": re.sub(r"\D", "", _text(row.get("zipCode"))),
                "area": area,
                "bedrooms": bedrooms,
                "bathrooms": bathrooms,
                "price": price,
            }
        )
        audit = {
            "observation_id": f"VIVAREAL-RJ-{index:04d}",
            "source_portal": "VIVAREAL",
            "source_url": _text(row.get("url")),
            "source_listing_id": _text(row.get("id")),
            "source_files": row.get("_source_files", ""),
            "captured_at": _text(row.get("scrapedAt")),
            "source_reference_date": _text(row.get("updatedAt")),
            "evidence_type": "OFFER",
            "property_type": PROPERTY_TYPE,
            "state": STATE,
            "city": CITY,
            "city_ibge_code": CITY_IBGE_CODE,
            "postal_code": re.sub(r"\D", "", _text(row.get("zipCode"))),
            "neighborhood": _text(row.get("neighborhood")),
            "street": _text(row.get("street")),
            "private_area_m2": area,
            "bedrooms": bedrooms,
            "bathrooms": bathrooms,
            "parking_spaces": parking,
            "asking_price_brl": price,
            "price_per_m2_brl": price_per_m2,
            "latitude_recovered": latitude,
            "longitude_recovered": longitude,
            "location_accuracy_meters": "",
            "physical_fingerprint": fingerprint,
            "training_eligible": False,
            "exclusion_reasons": reasons,
            "anomaly_flags": anomalies,
        }
        audit_rows.append(audit)

    chosen_by_fingerprint: dict[str, tuple[str, int]] = {}
    for index, audit in enumerate(audit_rows):
        if audit["exclusion_reasons"]:
            continue
        key = str(audit["physical_fingerprint"])
        rank = (str(audit["source_reference_date"]), index)
        previous = chosen_by_fingerprint.get(key)
        if previous is None or rank > previous:
            if previous is not None:
                audit_rows[previous[1]]["exclusion_reasons"].append(
                    "DUPLICATE_PHYSICAL_LISTING"
                )
            chosen_by_fingerprint[key] = rank
        else:
            audit["exclusion_reasons"].append("DUPLICATE_PHYSICAL_LISTING")
    for audit in audit_rows:
        audit["training_eligible"] = not audit["exclusion_reasons"]
        audit["exclusion_reasons"] = "|".join(audit["exclusion_reasons"])
        audit["anomaly_flags"] = "|".join(audit["anomaly_flags"])
    return audit_rows


def _write_audit_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as destination:
        writer = csv.DictWriter(destination, fieldnames=list(rows[0]))
        writer.writeheader()
        writer.writerows(rows)


def _admin_key(env_file: Path, requested_actor: str | None) -> tuple[str, str]:
    values = dotenv_values(env_file)
    raw = values.get("ADMIN_CREDENTIALS_JSON")
    if not raw:
        raise ValueError(f"{env_file}: ADMIN_CREDENTIALS_JSON ausente.")
    credentials = json.loads(raw)
    if not isinstance(credentials, dict) or not credentials:
        raise ValueError("ADMIN_CREDENTIALS_JSON deve conter ao menos uma identidade.")
    actor = requested_actor or sorted(credentials)[0]
    key = credentials.get(actor)
    if not isinstance(key, str) or not key:
        raise ValueError(f"Credencial administrativa não encontrada para {actor}.")
    return actor, key


def _http_json(
    *, base_url: str, path: str, key: str, method: str, payload: object | None = None
) -> dict[str, Any]:
    body = None
    headers = {"X-Admin-API-Key": key, "Accept": "application/json"}
    if payload is not None:
        body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        headers["Content-Type"] = "application/json; charset=utf-8"
    request = Request(
        f"{base_url.rstrip('/')}{path}", data=body, headers=headers, method=method
    )
    try:
        with urlopen(request, timeout=60) as response:
            return json.loads(response.read().decode("utf-8"))
    except HTTPError as error:
        detail = error.read().decode("utf-8", errors="replace")
        raise RuntimeError(f"API respondeu HTTP {error.code}: {detail}") from error
    except URLError as error:
        raise RuntimeError(f"API indisponível em {base_url}: {error.reason}") from error


def _download_report(*, base_url: str, model_id: str, key: str, path: Path) -> str:
    request = Request(
        f"{base_url.rstrip('/')}/statistical-models/{model_id}/report.pdf",
        headers={"X-Admin-API-Key": key, "Accept": "application/pdf"},
    )
    try:
        with urlopen(request, timeout=60) as response:
            content = response.read()
    except HTTPError as error:
        detail = error.read().decode("utf-8", errors="replace")
        raise RuntimeError(
            f"Relatório respondeu HTTP {error.code}: {detail}"
        ) from error
    if not content.startswith(b"%PDF-"):
        raise RuntimeError("A resposta do relatório não é um PDF válido.")
    path.write_bytes(content)
    return hashlib.sha256(content).hexdigest()


def _latest_reference_date(rows: Iterable[dict[str, Any]]) -> str:
    dates: list[str] = []
    for row in rows:
        value = str(row.get("source_reference_date", ""))
        if value:
            dates.append(value[:10])
    return max(dates) if dates else datetime.now(timezone.utc).date().isoformat()


def main() -> int:
    arguments = parse_arguments()
    paths = [path.resolve() for path in arguments.workbooks]
    if len(paths) != len(set(paths)):
        raise ValueError("Não informe o mesmo arquivo mais de uma vez.")
    for path in paths:
        if not path.is_file():
            raise FileNotFoundError(path)

    source_hashes = {path.name: file_sha256(path) for path in paths}
    loaded = [(path, load_rows(path)) for path in paths]
    reconciled, conflicts = reconcile_workbooks(loaded)
    audit_rows = prepare_audit_rows(reconciled)
    eligible = [row for row in audit_rows if row["training_eligible"]]
    if len(eligible) < 3:
        raise ValueError("Menos de três observações completas para diagnóstico.")

    output_dir = arguments.output_dir.resolve()
    output_dir.mkdir(parents=True, exist_ok=True)
    audit_csv = output_dir / "vivareal-rj-auditado.csv"
    manifest_path = output_dir / "vivareal-rj-manifest.json"
    model_json = output_dir / "vivareal-rj-modelo-pesquisa.json"
    report_pdf = output_dir / "RELATORIO-MODELO-VIVAREAL-RJ-PESQUISA.pdf"
    _write_audit_csv(audit_csv, audit_rows)

    training_rows = [
        [float(row[target_name]) for _, target_name in FEATURE_MAPPING]
        for row in eligible
    ]
    feature_names = [target for _, target in FEATURE_MAPPING]
    values = [float(row["asking_price_brl"]) for row in eligible]
    target = [median(column) for column in zip(*training_rows, strict=True)]
    reference_date = _latest_reference_date(eligible)
    source_fingerprint = _canonical_hash(
        {
            "source_hashes": source_hashes,
            "training_observations": [
                {
                    "source_url": row["source_url"],
                    "features": training_rows[index],
                    "asking_price_brl": values[index],
                }
                for index, row in enumerate(eligible)
            ],
        }
    )
    global_blockers = [
        "ASKING_PRICE_NOT_RT_ADJUSTED",
        "SOURCE_CONCENTRATION_100_PERCENT_VIVAREAL",
        "LOCATION_ACCURACY_NOT_DECLARED",
        "EVIDENCE_ARCHIVE_SHA256_NOT_PROVIDED",
        "GEOGRAPHIC_CONCENTRATION_COPACABANA",
    ]
    exclusion_counts = Counter(
        reason
        for row in audit_rows
        for reason in str(row["exclusion_reasons"]).split("|")
        if reason
    )
    anomaly_counts = Counter(
        reason
        for row in audit_rows
        for reason in str(row["anomaly_flags"]).split("|")
        if reason
    )
    neighborhood_counts = Counter(str(row["neighborhood"]) for row in eligible)
    manifest: dict[str, Any] = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "source_files": [
            {
                "path": str(path),
                "sha256": source_hashes[path.name],
                "row_count": len(rows),
            }
            for path, rows in loaded
        ],
        "reconciled_unique_urls": len(reconciled),
        "core_field_conflicts": conflicts,
        "audit_row_count": len(audit_rows),
        "research_training_row_count": len(eligible),
        "excluded_row_count": len(audit_rows) - len(eligible),
        "exclusion_reason_counts": dict(sorted(exclusion_counts.items())),
        "anomaly_counts": dict(sorted(anomaly_counts.items())),
        "training_neighborhood_counts": dict(sorted(neighborhood_counts.items())),
        "source_share": {"VIVAREAL": 1.0},
        "dependent_variable": "asking_price_brl",
        "training_classification": "RESEARCH_ONLY",
        "market_dataset_model_ready": False,
        "homologation_approval_allowed": False,
        "global_blockers": global_blockers,
        "source_fingerprint_sha256": source_fingerprint,
        "audit_csv": str(audit_csv),
    }

    if not arguments.skip_api:
        actor, key = _admin_key(arguments.env_file.resolve(), arguments.trainer_actor)
        timestamp = datetime.now(timezone.utc).strftime("%Y%m%d%H%M%S")
        payload = {
            "city_ibge_code": CITY_IBGE_CODE,
            "property_type": PROPERTY_TYPE,
            "dataset_version": f"VIVAREAL-RJ-{reference_date}-{source_fingerprint[:8]}",
            "source_reference": (
                f"{len(paths)} exportação(ões) VivaReal conciliada(s) e "
                f"deduplicada(s); lineage_sha256={source_fingerprint}"
            ),
            "reference_date": reference_date,
            "model_version": f"VIVAREAL-ASKING-OLS-{timestamp}",
            "valid_from": reference_date,
            "valid_until": None,
            "dataset_metadata": {
                "training_classification": "RESEARCH_ONLY",
                "market_dataset_model_ready": False,
                "source_portal": "VIVAREAL",
                "global_blockers": "|".join(global_blockers),
                "lineage_sha256": source_fingerprint,
            },
            "dependent_variable": "asking_price_brl",
            "dependent_variable_unit": "BRL",
            "dependent_variable_transformation": "NONE",
            "feature_transformations": {},
            "feature_names": feature_names,
            "observations": training_rows,
            "values": values,
            "target": target,
            "expected_signs": {name: 1 for name in feature_names},
            "confidence_level": 0.8,
        }
        model = _http_json(
            base_url=arguments.base_url,
            path="/statistical-models/train",
            key=key,
            method="POST",
            payload=payload,
        )
        if model.get("status") != "CANDIDATE":
            raise RuntimeError("Modelo de pesquisa não permaneceu como CANDIDATE.")
        model_json.write_text(
            json.dumps(model, ensure_ascii=False, indent=2) + "\n",
            encoding="utf-8",
        )
        report_hash = _download_report(
            base_url=arguments.base_url,
            model_id=str(model["model_id"]),
            key=key,
            path=report_pdf,
        )
        manifest["api_registration"] = {
            "base_url": arguments.base_url,
            "trainer_actor": actor,
            "model_id": model["model_id"],
            "dataset_id": model["dataset_id"],
            "model_status": model["status"],
            "dataset_status": model["dataset_status"],
            "model_artifact_sha256": model["artifact_sha256"],
            "report_pdf": str(report_pdf),
            "report_sha256": report_hash,
        }

    manifest_path.write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    print("Fonte(s) VivaReal conciliada(s) com rastreabilidade.")
    print(f"URLs únicas: {len(reconciled)}")
    print(f"Linhas usadas no treino exploratório: {len(eligible)}")
    print("Pronto para homologação: NÃO")
    print(f"CSV auditado: {audit_csv}")
    print(f"Manifesto: {manifest_path}")
    if not arguments.skip_api:
        print(f"Modelo de pesquisa: {model_json}")
        print(f"Relatório do modelo: {report_pdf}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
